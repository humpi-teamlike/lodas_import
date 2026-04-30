"""
LODAS Importskript  lodas_import_v2
=====================================
Liest Zeitnachweis-Excel-Dateien aus einem Monatsordner und erzeugt
eine LODAS-ASCII-Importdatei.

lodas_import_v2 – Änderungen gegenüber v1 (April 2026):
  - parse_datum(): Leitsystem-Datumsstrings 'DD.MM.YYYY Wt.' werden erkannt
  - lese_urlaubstage_aus_abwesenheiten(): urlaub_tage korrekt in v1 eingebaut
  - verarbeite_zeitnachweis_v1(): Schritt 4 urlaub_tage ergänzt

lodas_import_v1 – Neue Berechnungslogik (Basis):
  - Stunden werden 1:1 aus dem Zeitnachweis-Sheet uebernommen (keine 80%-Mechanik mehr)
  - LA 100: Summe aller Zeilen ohne Kommentar (geleistete Arbeitsstunden)
  - LA 202, 301, 320 etc.: Stunden direkt aus Zeitnachweis-Zeilen mit Kommentar
  - LA 123 (Freizeitausgleich): Tage x Tages-SOLL aus Abwesenheiten-Sheet
  - LA 122 / LA 124 (AZK): aus Saldo-Differenz der Uebersicht im Zeitnachweis-Sheet
  - Fehlzeiten (Schluessel 103, 200, 22): aus Abwesenheiten-Sheet
  - Dateisuche: Name aus Zelle C1 statt Dateiname (robust gegen Sonderzeichen-Bug)

Aufruf:
  python lodas_import_v1.py
"""

VERSION = "lodas_import_v2"

# ===========================================================================
# BLOCK 1 – Imports & Logging
# ===========================================================================
import os
import re
import logging
from collections import defaultdict
from datetime import datetime, timedelta, date

import pandas as pd
import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ===========================================================================
# BLOCK 2 – Konfiguration & Konstanten
# ===========================================================================
STAMMDATEN_PFAD = r"C:\Users\eikeh\01_ImportAPI\stammdaten.xlsx"

# Mapping: Kommentar im Zeitnachweis-Sheet (lowercase) → LODAS Lohnart-Nr.
KOMMENTAR_LOHNART_MAP = {
    "urlaub":       "301",
    "krank":        "320",
    "garantiezeit": "202",
    # Freizeitausgleich kommt aus Abwesenheiten-Sheet, nicht aus Zeitnachweis-Zeilen
}

# Mapping: Art im Abwesenheiten-Sheet (lowercase) → LODAS Fehlzeit-Schluessel
FEHLZEIT_SCHLUESSEL_MAP = {
    "krank":               "103",
    "kind krank":          "200",
    "unbezahlte fehlzeit": "22",
}

# parse_datum: Leitsystem exportiert Datum als String 'DD.MM.YYYY Wt.' statt datetime
def parse_datum(wert):
    """Erkennt datetime/date-Objekt ODER Leitsystem-String 'DD.MM.YYYY Wt.' (z. B. '02.03.2026 Mo.').
    Rückgabe: date-Objekt oder None.
    FIX v2: Das Leitsystem exportiert Datumswerte als String, nicht als datetime-Objekt.
    """
    if isinstance(wert, (datetime, date)):
        return wert
    if isinstance(wert, str):
        try:
            return datetime.strptime(wert.strip()[:10], "%d.%m.%Y").date()
        except ValueError:
            return None
    return None


# ===========================================================================
# BLOCK 3 – Hilfsfunktionen
# ===========================================================================
def safe_float(val, default=0.0):
    try:
        f = float(val)
        return default if pd.isna(f) else f
    except (TypeError, ValueError):
        return default

def clean(val, none_if_empty=False):
    s = str(val or "").strip().replace("\r", "").replace("\n", "")
    empty = None if none_if_empty else ""
    return empty if s in ("", "nan") else s

def norm_col(col):
    return re.sub(r"[\r\n\s]+", " ", str(col)).strip()

def parse_mmyyyy(mmyyyy_str):
    mmyyyy_str = mmyyyy_str.strip()
    if not re.match(r"^\d{6}$", mmyyyy_str):
        raise ValueError(f"Ungaeltiges Format '{mmyyyy_str}' - erwartet: MMYYYY (z. B. 042026)")
    mm = mmyyyy_str[:2]
    yyyy = mmyyyy_str[2:]
    if not (1 <= int(mm) <= 12):
        raise ValueError(f"Ungueltiger Monat '{mm}'")
    return f"01.{mm}.{yyyy}"

def ask(prompt, default=None):
    if default:
        val = input(f"{prompt} [{default}]: ").strip()
        return val if val else default
    return input(f"{prompt}: ").strip()

def schreibe_datei(pfad, zeilen, encoding):
    inhalt = "\r\n".join(z.replace("\r", "").replace("\n", "") for z in zeilen)
    with open(pfad, "w", encoding=encoding, errors="replace", newline="") as f:
        f.write(inhalt)

# ===========================================================================
# BLOCK 4 – ReportCollector
# ===========================================================================
class ReportCollector:
    SEP  = "=" * 70
    SEP2 = "-" * 70

    def __init__(self):
        self.mitarbeiter   = []
        self.fehler        = []
        self.uebersprungen = []

    def add_ma(self, name, pnr, la100_h, azk_la, azk_h, fehlzeiten_anzahl,
               saldo_differenz=None, fza_stunden=0.0):
        self.mitarbeiter.append({
            "name":              name,
            "pnr":               pnr,
            "la100_h":           la100_h,
            "azk_la":            azk_la,
            "azk_h":             azk_h,
            "fehlzeiten_anzahl": fehlzeiten_anzahl,
            "saldo_differenz":   saldo_differenz,
            "fza_stunden":       fza_stunden,
        })

    def hinweis(self, ma, text):
        self.fehler.append({"ebene": "HINWEIS", "ma": ma, "text": text})

    def fehler_add(self, ma, text):
        self.fehler.append({"ebene": "FEHLER", "ma": ma, "text": text})

    def skip(self, dateiname, grund):
        self.uebersprungen.append(dateiname)
        self.fehler.append({"ebene": "FEHLER", "ma": dateiname, "text": grund})

    def add_error(self, ma, text):
        """Alias für fehler_add – für BLOCK 7 Kompatibilität."""
        self.fehler.append({"ebene": "FEHLER", "ma": ma, "text": text})

    def add_warning(self, ma, text):
        """Warnung loggen – erscheint im Report als HINWEIS."""
        self.fehler.append({"ebene": "HINWEIS", "ma": ma, "text": f"[WARNUNG] {text}"})

    def add_info(self, ma, text):
        """Info loggen – erscheint im Report als HINWEIS."""
        self.fehler.append({"ebene": "HINWEIS", "ma": ma, "text": f"[INFO] {text}"})

    def render(self, az_eingabe, az_anzeige, gesamt_bwd, gesamt_fz, encoding="ansi"):
        col_name=22; col_pnr=5; col_la100=10; col_azk=14; col_fz=12; col_hw=8
        zeilen = []
        zeilen.append(self.SEP)
        zeilen.append(f"  LODAS Import Report  [{VERSION}]  -  {az_eingabe}  ({az_anzeige})")
        zeilen.append(self.SEP)
        zeilen.append("")
        zeilen.append(
            f"  {'Mitarbeiter':<{col_name}} {'PNr':<{col_pnr}} "
            f"{'LA 100':>{col_la100}} {'AZK-Buchung':>{col_azk}} "
            f"{'Fehlzeiten':>{col_fz}} {'Hinweise':>{col_hw}}"
        )
        zeilen.append("  " + self.SEP2)
        for ma in self.mitarbeiter:
            hinweis_count = sum(1 for e in self.fehler if e["ma"] == ma["name"])
            fehler_count  = sum(1 for e in self.fehler if e["ma"] == ma["name"] and e["ebene"] == "FEHLER")
            la100_str = f"{ma['la100_h']:.2f}h" if ma["la100_h"] is not None else "-"
            if ma["azk_la"] == "122":   azk_str = f"LA122 -{ma['azk_h']:.2f}h"
            elif ma["azk_la"] == "124": azk_str = f"LA124 +{ma['azk_h']:.2f}h"
            else:                       azk_str = "-"
            fz_str = f"{ma['fehlzeiten_anzahl']}x Fehlzeit" if ma["fehlzeiten_anzahl"] > 0 else "-"
            if fehler_count > 0:    hw_str = f"{hinweis_count}x E"
            elif hinweis_count > 0: hw_str = f"{hinweis_count}x !"
            else:                   hw_str = "OK"
            zeilen.append(
                f"  {ma['name']:<{col_name}} {ma['pnr']:<{col_pnr}} "
                f"{la100_str:>{col_la100}} {azk_str:>{col_azk}} "
                f"{fz_str:>{col_fz}} {hw_str:>{col_hw}}"
            )
        if self.uebersprungen:
            zeilen.append("  " + self.SEP2)
            for _ in self.uebersprungen:
                zeilen.append(
                    f"  {'[uebersprungen]':<{col_name}} {'':>{col_pnr}} "
                    f"{'':>{col_la100}} {'':>{col_azk}} {'':>{col_fz}} {'FEHLER':>{col_hw}}"
                )
        zeilen.append("")
        zeilen.append(self.SEP)
        zeilen.append(f"  Verarbeitete Mitarbeiter:  {len(self.mitarbeiter)}")
        zeilen.append(f"  Uebersprungene Dateien:    {len(self.uebersprungen)}")
        zeilen.append(f"  Erzeugte Bewegungsdaten:   {gesamt_bwd}")
        zeilen.append(f"  Erzeugte Fehlzeiten:       {gesamt_fz}")
        zeilen.append(self.SEP)
        if self.fehler:
            zeilen.append("")
            zeilen.append("  HINWEISE & FEHLER")
            zeilen.append("  " + self.SEP2)
            for e in self.fehler:
                symbol = "!!" if e["ebene"] == "FEHLER" else " !"
                zeilen.append(f"  [{symbol}] [{e['ma']}]  {e['text']}")
            zeilen.append(self.SEP)
        else:
            zeilen.append("")
            zeilen.append("  Keine Hinweise oder Fehler - Verarbeitung fehlerfrei.")
            zeilen.append(self.SEP)
        return zeilen

    def render_sync(self, az_eingabe, az_anzeige):
        col_pnr=6; col_name=28; col_sal=10; col_azk=10; col_nach=11
        zeilen = []
        zeilen.append(self.SEP)
        zeilen.append(f"  LODAS Sync-Report  [{VERSION}]  -  {az_eingabe}  ({az_anzeige})")
        zeilen.append(self.SEP)
        zeilen.append(
            f"  {'PNr':<{col_pnr}} {'Name':<{col_name}} "
            f"{'Saldo LS':>{col_sal}} {'AZK Netto':>{col_azk}} "
            f"{'Nachbuchung':>{col_nach}}"
        )
        zeilen.append("  " + self.SEP2)
        for ma in self.mitarbeiter:
            saldo = ma["saldo_differenz"]
            if ma["azk_la"] == "122":
                azk_netto = -ma["azk_h"]
            elif ma["azk_la"] == "124":
                azk_netto = ma["azk_h"]
            else:
                azk_netto = 0.0
            azk_str = f"{azk_netto:+.2f}"
            if saldo is not None:
                saldo_str       = f"{saldo:+.2f}"
                nachbuchung_str = f"{round(saldo - azk_netto, 2):+.2f}"
            else:
                saldo_str       = "n/a"
                nachbuchung_str = "n/a"
            zeilen.append(
                f"  {ma['pnr']:<{col_pnr}} {ma['name']:<{col_name}} "
                f"{saldo_str:>{col_sal}} {azk_str:>{col_azk}} "
                f"{nachbuchung_str:>{col_nach}}"
            )
        zeilen.append(self.SEP)
        return zeilen


report = ReportCollector()

# ===========================================================================
# BLOCK 5 – Stammdaten einlesen
# ===========================================================================
print("=" * 60)
print(f"  LODAS Importskript  {VERSION}")
print("=" * 60)
print(f"  Stammdaten: {STAMMDATEN_PFAD}")
print("=" * 60)

while True:
    az_eingabe = ask("Abrechnungszeitraum (Format MMYYYY, z. B. 042026)")
    try:
        abrechnungszeitraum_global = parse_mmyyyy(az_eingabe)
        log.info(f"Abrechnungszeitraum: {abrechnungszeitraum_global}")
        break
    except ValueError as e:
        print(f"  Fehler: {e} - bitte erneut eingeben.")

zeitnachweis_ordner = rf"C:\Users\eikeh\01_ImportAPI\Lohnexport\{az_eingabe}"
log.info(f"Zeitnachweis-Ordner: {zeitnachweis_ordner}")

def read_sheet_find_header(pfad, sheet, schluessel_spalte):
    raw = pd.read_excel(pfad, sheet_name=sheet, header=None)
    for i, row in raw.iterrows():
        if any(schluessel_spalte in str(v) for v in row.values):
            raw.columns = [norm_col(c) for c in raw.iloc[i]]
            df = raw.iloc[i + 1:].reset_index(drop=True)
            return df
    raise ValueError(f"Spalte mit '{schluessel_spalte}' nicht im Sheet '{sheet}' gefunden.")

log.info("Lese Stammdaten...")

if not os.path.exists(STAMMDATEN_PFAD):
    log.error(f"stammdaten.xlsx nicht gefunden: {STAMMDATEN_PFAD}")
    exit(1)

raw_mandant = pd.read_excel(STAMMDATEN_PFAD, sheet_name="Mandant", header=None)
mandant = {}
header_found = False
for i, row in raw_mandant.iterrows():
    vals = [str(v).strip() for v in row.values]
    if "Feld" in vals and not header_found:
        header_found = True
        continue
    if header_found:
        k = clean(row.iloc[0])
        v = clean(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
        if k:
            mandant[k] = v

df_mitarbeiter = read_sheet_find_header(STAMMDATEN_PFAD, "Mitarbeiter", "Name (wie in Zeitnachweis)")
df_mitarbeiter = df_mitarbeiter.dropna(subset=["Name (wie in Zeitnachweis)"])

mitarbeiter_liste = []
for _, row in df_mitarbeiter.iterrows():
    name = clean(row["Name (wie in Zeitnachweis)"])
    if not name:
        continue
    pnr = clean(row.get("Personalnummer", "")) or ""
    if not re.match(r"^\d{1,3}$", pnr):
        log.warning(f"Personalnummer '{pnr}' fuer '{name}' ignoriert")
        continue
    bemerkung  = clean(row.get("Bemerkung", "") or "")
    ist_gehalt = bemerkung.lower() == "gehalt"
    mitarbeiter_liste.append({
        "name":              name,
        "ma_nr":             pnr,
        "tages_soll":        safe_float(row.get("Tages Arbeitszeit SOLL")),
        "tages_ist":         safe_float(row.get("Tages Arbeitszeit IST")),
        "auszahlungsgrenze": safe_float(row.get("Auszahlungsgrenze", 0)),
        "ist_gehalt":        ist_gehalt,
    })

encoding = mandant.get("Encoding", "ansi")
log.info(f"  {len(mitarbeiter_liste)} Mitarbeiter geladen")

# ===========================================================================
# BLOCK 6 – Zeitnachweis verarbeiten (lodas_import_v1 – NEU)
# ===========================================================================

def lese_stunden_aus_zeitnachweis(ws_zeitnachweis, mitarbeiter_name):
    """
    Liest alle Stunden-Zeilen aus dem Zeitnachweis-Sheet.
    Spalte A (Index 0) = Datum, Spalte E (Index 4) = Stunden, Spalte G (Index 6) = Kommentar.
    Leere Kommentare → LA 100 (Stundenlohn).
    Bekannte Kommentare → gemaess KOMMENTAR_LOHNART_MAP.
    Unbekannte Kommentare → Warnung im Report.
    Gibt zurueck: { lohnart_nr: stunden_float }
    """
    stunden_je_lohnart  = {}
    unbekannte_kommentare = []
    vorschuesse         = []

    for row in ws_zeitnachweis.iter_rows(values_only=True):
        if not row or len(row) < 5:
            continue

        datum_zelle     = row[0]   # Spalte A – Datum
        stunden_zelle   = row[4]   # Spalte E – Stunden
        kommentar_zelle = row[6] if len(row) > 6 else None  # Spalte G – Kommentar

        # Datum erkennen: datetime/date-Objekt ODER Leitsystem-String 'DD.MM.YYYY Wt.' (FIX v2)
        datum = parse_datum(datum_zelle)
        if datum is None:
            continue

        try:
            stunden = float(stunden_zelle) if stunden_zelle not in (None, "") else 0.0
        except (TypeError, ValueError):
            stunden = 0.0

        if stunden == 0.0:
            continue

        kommentar = str(kommentar_zelle).strip() if kommentar_zelle else ""
        kommentar_lower = kommentar.lower()

        # Abschlag/Vorschuss separat sammeln
        if kommentar_lower.startswith("abschlag") or kommentar_lower.startswith("vorschuss"):
            teile = kommentar.split(None, 1)
            if len(teile) == 2:
                betrag = safe_float(teile[1].replace(",", "."))
                if betrag > 0:
                    vorschuesse.append({
                        "pnr":                 None,  # wird in BLOCK 7 gesetzt
                        "abrechnungszeitraum": abrechnungszeitraum_global,
                        "betrag":              betrag,
                    })
                else:
                    report.hinweis(mitarbeiter_name, f"Abschlag ohne gueltigen Betrag: '{kommentar}'")
            else:
                report.hinweis(mitarbeiter_name, f"Abschlag-Kommentar ohne Betrag: '{kommentar}'")
            continue

        # Lohnart bestimmen
        if kommentar == "" or kommentar_lower == "none":
            lohnart = "100"
        elif kommentar_lower in KOMMENTAR_LOHNART_MAP:
            lohnart = KOMMENTAR_LOHNART_MAP[kommentar_lower]
        else:
            if kommentar_lower not in unbekannte_kommentare:
                unbekannte_kommentare.append(kommentar_lower)
            continue

        stunden_je_lohnart[lohnart] = round(
            stunden_je_lohnart.get(lohnart, 0.0) + stunden, 2
        )

    for k in unbekannte_kommentare:
        report.hinweis(mitarbeiter_name, f"Unbekannter Kommentar '{k}' – keine Lohnart zugeordnet, nicht gebucht.")

    return stunden_je_lohnart, vorschuesse


def lese_fza_aus_abwesenheiten(ws_abwesenheiten, tages_soll, mitarbeiter_name):
    """
    Liest Freizeitausgleich-Eintraege aus dem Abwesenheiten-Sheet.
    Tage x Tages-SOLL = LA 123 Stunden (Auszahlung).
    Der AZK-Abzug ist bereits im Leitsystem-Saldo enthalten – kein separater Abzug noetig.
    """
    fza_stunden = 0.0
    fza_tage    = 0

    if ws_abwesenheiten is None:
        return 0.0

    header_gefunden = False
    for row in ws_abwesenheiten.iter_rows(values_only=True):
        if not row:
            continue
        # Header-Zeile erkennen (enthaelt "Beginn")
        if not header_gefunden:
            if any("Beginn" in str(v) for v in row if v):
                header_gefunden = True
            continue

        if len(row) < 3:
            continue

        von_zelle = row[0]
        bis_zelle = row[1]
        art_zelle = row[2]

        if not art_zelle:
            continue

        art = str(art_zelle).strip().lower()
        if art != "freizeitausgleich":
            continue

        try:
            if isinstance(von_zelle, (datetime, date)) and isinstance(bis_zelle, (datetime, date)):
                delta = (bis_zelle - von_zelle).days + 1
            else:
                delta = 1
        except Exception:
            delta = 1

        fza_tage    += delta
        fza_stunden += round(delta * tages_soll, 2)

    if fza_stunden > 0:
        report.hinweis(mitarbeiter_name,
            f"Freizeitausgleich: {fza_tage} Tag(e) x {tages_soll}h SOLL = {fza_stunden:.2f}h → LA 123")

    return round(fza_stunden, 2)


def lese_fehlzeiten_aus_abwesenheiten(ws_abwesenheiten, mitarbeiter_name):
    """
    Liest Fehlzeiten aus dem Abwesenheiten-Sheet.
    Gibt Liste von Dicts zurueck: { schluessel, von, bis, art }
    FZA und Urlaub werden hier nicht als Fehlzeit gebucht (Backlog #20).
    """
    fehlzeiten = []

    if ws_abwesenheiten is None:
        return fehlzeiten

    header_gefunden = False
    for row in ws_abwesenheiten.iter_rows(values_only=True):
        if not row:
            continue
        if not header_gefunden:
            if any("Beginn" in str(v) for v in row if v):
                header_gefunden = True
            continue

        if len(row) < 3:
            continue

        von_zelle = row[0]
        bis_zelle = row[1]
        art_zelle = row[2]

        if not art_zelle:
            continue

        art = str(art_zelle).strip().lower()

        if art not in FEHLZEIT_SCHLUESSEL_MAP:
            continue  # FZA, Urlaub etc. – Backlog #20

        schluessel = FEHLZEIT_SCHLUESSEL_MAP[art]

        try:
            von = von_zelle if isinstance(von_zelle, (datetime, date)) else None
            bis = bis_zelle if isinstance(bis_zelle, (datetime, date)) else None
            if not von or not bis:
                raise ValueError("kein gueltiges Datum")
        except Exception:
            report.hinweis(mitarbeiter_name,
                f"Fehlzeit '{art_zelle}' ohne gueltiges Datum – wird uebersprungen.")
            continue

        fehlzeiten.append({
            "schluessel": schluessel,
            "von":        von,
            "bis":        bis,
            "art":        art_zelle,
        })
        log.info(f"  Fehlzeit: {art_zelle} | "
                 f"{von.strftime('%d.%m.%Y')} – {bis.strftime('%d.%m.%Y')} | "
                 f"Schluessel {schluessel}")

    return fehlzeiten


def lese_azk_saldo_differenz(ws_zeitnachweis, mitarbeiter_name):
    """
    Liest die Saldo-Differenz aus der Uebersicht-Tabelle im Zeitnachweis-Sheet.
    Anker: Zeile mit Label 'Uebersicht' (positionsunabhaengig).
    Danach: Zeile mit Label 'Saldo' – erste zwei numerische Werte = Vormonat / aktueller Monat.
    Positiv → AZK gestiegen → LA 122 mit negativem Wert buchen.
    Negativ → AZK gesunken  → LA 124 mit positivem Wert buchen.
    """
    in_uebersicht   = False
    saldo_differenz = None

    for row in ws_zeitnachweis.iter_rows(values_only=True):
        if not row:
            continue

        label = str(row[0]).strip().lower() if row[0] else ""

        if not in_uebersicht:
            if label in ("übersicht", "ubersicht", "uebersicht"):
                in_uebersicht = True
            continue

        if "saldo" not in label:
            continue

        werte = [
            float(cell) for cell in row[1:]
            if cell is not None
            and str(cell).strip() not in ("", "None")
            and isinstance(cell, (int, float))
        ]

        if len(werte) >= 2:
            saldo_vormonat  = werte[0]
            saldo_aktuell   = werte[1]
            saldo_differenz = round(saldo_aktuell - saldo_vormonat, 2)
            break

    if saldo_differenz is None:
        report.hinweis(mitarbeiter_name,
            "Saldo-Differenz nicht gefunden – keine AZK-Buchung erzeugt.")
        return None

    if saldo_differenz > 0:
        log.info(f"  AZK gestiegen: +{saldo_differenz}h → LA 122 mit -{saldo_differenz}h")
    elif saldo_differenz < 0:
        log.info(f"  AZK gesunken: {saldo_differenz}h → LA 124 mit +{abs(saldo_differenz)}h")
    else:
        log.info("  Saldo unveraendert – keine AZK-Buchung.")

    return saldo_differenz


def lese_urlaubstage_aus_abwesenheiten(ws_abwesenheiten, mitarbeiter_name):
    """
    Zählt Urlaubs-Tage aus dem Abwesenheiten-Sheet (Art = 'Urlaub').
    Rückgabe: int (Anzahl Kalendertage)
    Wird für LA 303 (Urlaub, Tagebasis, bs_nr=71) verwendet.
    NEU in v2.
    """
    if ws_abwesenheiten is None:
        return 0

    urlaub_tage    = 0
    header_gefunden = False

    for row in ws_abwesenheiten.iter_rows(values_only=True):
        if not row:
            continue
        if not header_gefunden:
            if any("Beginn" in str(v) for v in row if v):
                header_gefunden = True
            continue

        if len(row) < 3:
            continue

        von_zelle = row[0]
        bis_zelle = row[1]
        art_zelle = row[2]

        if not art_zelle:
            continue

        art = str(art_zelle).strip().lower()
        if art != "urlaub":
            continue

        try:
            if isinstance(von_zelle, (datetime, date)) and isinstance(bis_zelle, (datetime, date)):
                delta = (bis_zelle - von_zelle).days + 1
            else:
                delta = 1
        except Exception:
            delta = 1

        urlaub_tage += delta

    if urlaub_tage > 0:
        log.info(f"  Urlaub: {urlaub_tage} Tag(e) aus Abwesenheiten → LA 303")

    return urlaub_tage


def verarbeite_zeitnachweis_v1(ws_zeitnachweis, ws_abwesenheiten, stammdaten, mitarbeiter_name):
    """
    Koordiniert BLOCK 6: liest Stunden, FZA, Fehlzeiten und AZK-Saldo.
    Gibt ein strukturiertes Ergebnis-Dict zurueck.
    """
    tages_soll = stammdaten.get("tages_soll", 0.0)

    # Schritt 1: Stunden direkt aus Zeitnachweis-Sheet
    lohnarten, vorschuesse = lese_stunden_aus_zeitnachweis(ws_zeitnachweis, mitarbeiter_name)

    # Schritt 2: FZA aus Abwesenheiten → LA 123
    fza_stunden = lese_fza_aus_abwesenheiten(ws_abwesenheiten, tages_soll, mitarbeiter_name)
    if fza_stunden > 0:
        lohnarten["123"] = round(lohnarten.get("123", 0.0) + fza_stunden, 2)

    # Schritt 3: Fehlzeiten aus Abwesenheiten
    fehlzeiten = lese_fehlzeiten_aus_abwesenheiten(ws_abwesenheiten, mitarbeiter_name)

    # Schritt 4: Urlaubstage aus Abwesenheiten → LA 303 (NEU v2)
    urlaub_tage = lese_urlaubstage_aus_abwesenheiten(ws_abwesenheiten, mitarbeiter_name)

    # Schritt 5: AZK-Saldo-Differenz aus Uebersicht im Zeitnachweis-Sheet
    azk_differenz = lese_azk_saldo_differenz(ws_zeitnachweis, mitarbeiter_name)

    log.info(f"  Lohnarten-Uebersicht: " +
             ", ".join(f"LA {la}={h}h" for la, h in sorted(lohnarten.items())))

    return {
        "lohnarten":     lohnarten,
        "fza_stunden":   fza_stunden,
        "azk_differenz": azk_differenz,
        "fehlzeiten":    fehlzeiten,
        "vorschuesse":   vorschuesse,
        "urlaub_tage":   urlaub_tage,   # NEU v2
    }

# ===========================================================================
# BLOCK 7 – Hauptverarbeitung (lodas_import_v1)
# ===========================================================================

def baue_bewegungsdaten_zeilen(ma_nr, abrechnungsmonat, lohnarten_dict):
    """Wandelt lohnarten-Dict in LODAS-Bewegungsdaten-Zeilen um."""
    zeilen = []
    for lohnart, stunden in sorted(lohnarten_dict.items()):
        if stunden == 0.0:
            continue
        zeilen.append({
            "typ":     "L",
            "ma_nr":   ma_nr,
            "lohnart": lohnart,
            "stunden": stunden,
            "monat":   abrechnungsmonat,
        })
    return zeilen


def baue_azk_zeile(ma_nr, abrechnungsmonat, azk_differenz):
    """
    Positiv → LA 122 (negativ in LODAS, AZK aufgebaut)
    Negativ → LA 124 (positiv in LODAS, AZK abgebaut)
    """
    if azk_differenz is None or azk_differenz == 0.0:
        return None
    if azk_differenz > 0:
        return {"typ": "L", "ma_nr": ma_nr, "lohnart": "122",
                "stunden": round(-azk_differenz, 2), "monat": abrechnungsmonat}
    else:
        return {"typ": "L", "ma_nr": ma_nr, "lohnart": "124",
                "stunden": round(abs(azk_differenz), 2), "monat": abrechnungsmonat}


def baue_fehlzeit_zeilen(ma_nr, fehlzeiten):
    """Wandelt Fehlzeit-Einträge in LODAS-Fehlzeit-Zeilen um."""
    zeilen = []
    for f in fehlzeiten:
        zeilen.append({
            "typ":        "F",
            "ma_nr":      ma_nr,
            "schluessel": f["schluessel"],
            "von":        f["von"].strftime("%d.%m.%Y") if hasattr(f["von"], "strftime") else str(f["von"]),
            "bis":        f["bis"].strftime("%d.%m.%Y") if hasattr(f["bis"], "strftime") else str(f["bis"]),
            "art":        f["art"],
        })
    return zeilen


def verarbeite_mitarbeiter_v1(zeitnachweis_pfad, stammdaten_row, abrechnungsmonat, report):
    """
    Verarbeitet einen Mitarbeiter vollständig.
    stammdaten_row: { ma_nr, name, tages_soll, tages_ist, ist_gehalt, auszahlungsgrenze }
    """
    import openpyxl

    ma_nr      = stammdaten_row.get("ma_nr")
    name       = stammdaten_row.get("name", f"MA {ma_nr}")
    tages_soll = float(stammdaten_row.get("tages_soll", 0.0))
    ist_gehalt = bool(stammdaten_row.get("ist_gehalt", False))

    stammdaten = {
        "tages_soll":        tages_soll,
        "tages_ist":         float(stammdaten_row.get("tages_ist", 0.0)),
        "auszahlungsgrenze": stammdaten_row.get("auszahlungsgrenze"),
    }

    try:
        wb = openpyxl.load_workbook(zeitnachweis_pfad, data_only=True)
    except Exception as e:
        report.add_error(name, f"Zeitnachweis konnte nicht geöffnet werden: {e}")
        return None

    ws_zeitnachweis = None
    for sheet_name in wb.sheetnames:
        if "zeitnachweis" in sheet_name.lower():
            ws_zeitnachweis = wb[sheet_name]
            break
    if ws_zeitnachweis is None:
        ws_zeitnachweis = wb.worksheets[0]
        report.add_warning(name, f"Kein Sheet 'Zeitnachweis' – nutze erstes Sheet: '{ws_zeitnachweis.title}'")

    ws_abwesenheiten = None
    for sheet_name in wb.sheetnames:
        if "abwesenheit" in sheet_name.lower():
            ws_abwesenheiten = wb[sheet_name]
            break
    if ws_abwesenheiten is None:
        report.add_warning(name, "Kein Abwesenheiten-Sheet – FZA, Fehlzeiten und LA 303 werden übersprungen.")

    ergebnis = verarbeite_zeitnachweis_v1(
        ws_zeitnachweis  = ws_zeitnachweis,
        ws_abwesenheiten = ws_abwesenheiten,
        stammdaten       = stammdaten,
        mitarbeiter_name = name,
        report           = report,
    )

    if ergebnis is None:
        report.add_error(name, "BLOCK 6 hat kein Ergebnis – Mitarbeiter wird übersprungen.")
        return None

    ausgabe = {
        "ma_nr":      ma_nr,
        "name":       name,
        "bewegung":   [],
        "fehlzeiten": [],
        "azk":        None,
    }

    urlaub_tage = ergebnis.get("urlaub_tage", 0)
    lohnarten   = ergebnis["lohnarten"]

    if ist_gehalt:
        # ── GEHALTSEMPFÄNGER: nur LA 303 (Urlaub, Tagebasis, bs_nr=71) ──
        report.add_info(name, "Gehaltsempfänger → nur LA 303 wird gebucht.")
        if urlaub_tage > 0:
            ausgabe["bewegung"].append({
                "typ":     "L",
                "ma_nr":   ma_nr,
                "lohnart": "303",
                "stunden": urlaub_tage,
                "bs_nr":   71,
                "monat":   abrechnungsmonat,
            })
        else:
            report.add_info(name, "Kein Urlaub in Abwesenheiten → keine LA 303-Buchung.")

    else:
        # ── LOHNEMPFÄNGER: LA 301 (Zeitnachweis) + LA 303 (Abwesenheiten) + Rest ──
        report.add_info(name, "Lohnempfänger → LA 301 + LA 303 + AZK + Fehlzeiten.")

        ausgabe["bewegung"] = baue_bewegungsdaten_zeilen(
            ma_nr, abrechnungsmonat, lohnarten
        )

        if urlaub_tage > 0:
            ausgabe["bewegung"].append({
                "typ":     "L",
                "ma_nr":   ma_nr,
                "lohnart": "303",
                "stunden": urlaub_tage,
                "bs_nr":   71,
                "monat":   abrechnungsmonat,
            })

        azk_zeile = baue_azk_zeile(ma_nr, abrechnungsmonat, ergebnis["azk_differenz"])
        if azk_zeile:
            ausgabe["azk"] = azk_zeile
            ausgabe["bewegung"].append(azk_zeile)

        ausgabe["fehlzeiten"] = baue_fehlzeit_zeilen(ma_nr, ergebnis["fehlzeiten"])

    report.add_info(name, "── Ergebnis BLOCK 7 ──")
    report.add_info(name, f"  Typ:             {'Gehalt' if ist_gehalt else 'Lohn'}")
    report.add_info(name, f"  Bewegungsdaten:  {len(ausgabe['bewegung'])} Zeile(n)")
    report.add_info(name, f"  Fehlzeiten:      {len(ausgabe['fehlzeiten'])} Eintrag(e)")
    if not ist_gehalt:
        azk_info = ausgabe.get("azk")
        report.add_info(name, f"  AZK-Buchung:     {'ja – ' + str(azk_info) if azk_info else 'keine'}")

    return ausgabe


def name_aus_xlsx(pfad):
    """Liest Mitarbeiternamen aus C1 des Zeitnachweis-Sheets."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(pfad, data_only=True, read_only=True)
        ws = None
        for sheet_name in wb.sheetnames:
            if "zeitnachweis" in sheet_name.lower():
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.worksheets[0]
        name_raw = ws.cell(row=1, column=3).value
        wb.close()
        if not name_raw:
            return None
        teile = str(name_raw).strip().split(",", 1)
        if len(teile) == 2:
            return f"{teile[0].strip()}, {teile[1].strip()}"
        return str(name_raw).strip()
    except Exception:
        return None


def baue_datei_index(zeitnachweis_dir, report):
    """Erstellt { 'Nachname, Vorname' : pfad } aus xlsx-Dateien im Ordner."""
    import os
    index = {}
    for dateiname in os.listdir(zeitnachweis_dir):
        if not dateiname.lower().endswith(".xlsx"):
            continue
        pfad = os.path.join(zeitnachweis_dir, dateiname)
        name = name_aus_xlsx(pfad)
        if name:
            if name in index:
                report.add_warning(name,
                    f"Doppelte Datei: '{dateiname}' – bereits '{os.path.basename(index[name])}'. Bitte prüfen.")
            else:
                index[name] = pfad
        else:
            report.add_warning(dateiname,
                f"Name nicht aus C1 lesbar – Datei übersprungen: '{dateiname}'")
    return index


def hauptverarbeitung_v1(config, stammdaten_liste, report):
    """
    Schleife über alle Mitarbeiter aus der Stammdaten-Liste.
    config: { abrechnungsmonat, zeitnachweis_ordner }
    """
    import os
    abrechnungsmonat = config["abrechnungsmonat"]
    zeitnachweis_dir = config["zeitnachweis_ordner"]

    report.add_info("SYSTEM", f"Lese Zeitnachweis-Ordner: {zeitnachweis_dir}")
    datei_index = baue_datei_index(zeitnachweis_dir, report)
    report.add_info("SYSTEM", f"{len(datei_index)} Zeitnachweis-Datei(en) indexiert.")

    alle_ergebnisse = []

    for row in stammdaten_liste:
        ma_nr = row.get("ma_nr")
        name  = row.get("name", f"MA {ma_nr}")

        zeitnachweis_pfad = datei_index.get(name)
        if not zeitnachweis_pfad:
            report.add_error(name,
                f"Kein Zeitnachweis ('{name}' nicht im Ordner). "
                f"Verfügbar: {list(datei_index.keys())}")
            continue

        report.add_info(name, f"Zeitnachweis: {os.path.basename(zeitnachweis_pfad)}")

        ergebnis = verarbeite_mitarbeiter_v1(
            zeitnachweis_pfad = zeitnachweis_pfad,
            stammdaten_row    = row,
            abrechnungsmonat  = abrechnungsmonat,
            report            = report,
        )
        if ergebnis:
            alle_ergebnisse.append(ergebnis)

    report.add_info("GESAMT",
        f"{len(alle_ergebnisse)} von {len(stammdaten_liste)} Mitarbeiter(n) erfolgreich verarbeitet.")
    return alle_ergebnisse

# Ende BLOCK 7
